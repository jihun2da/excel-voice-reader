import streamlit as st
import openpyxl
import tempfile
import os
import uuid
import asyncio
import edge_tts
import io
import base64
from datetime import datetime
import re
import pandas as pd

# 페이지 설정
st.set_page_config(
    page_title="🎵 엑셀 음성 리더",
    page_icon="🎵",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 사이즈 변환 딕셔너리
SIZE_DICT = {
    "XS": "엑스스몰", "S": "스몰", "M": "미디움", "L": "라지", "FREE": "프리",
    "XL": "엑스라지", "XXL": "투엑스라지",
    "JS": "주니어 스몰", "JM": "주니어 미디움", "JL": "주니어 라지"
}

# 한국어 숫자 변환
KOREAN_NUMBER_MAP = {
    1: "한개", 2: "두개", 3: "세개", 4: "네개",
    5: "다섯개", 6: "여섯개", 7: "일곱개",
    8: "여덟개", 9: "아홉개", 10: "열개"
}

def convert_size(size_code):
    return SIZE_DICT.get(str(size_code).upper(), size_code)

def convert_quantity(n):
    if isinstance(n, int) and n >= 1:
        return KOREAN_NUMBER_MAP.get(n, f"{n}개")
    return ""

def clean_g_value(text):
    return re.sub(r"\(.*?\)", "", text).strip()

def count_consecutive_g_values(ws, start_row):
    current_g = clean_g_value(str(ws.cell(row=start_row, column=7).value or ""))
    count = 0
    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        g_value = clean_g_value(str(ws.cell(row=row, column=7).value or ""))
        if g_value == current_g:
            count += 1
        else:
            break
    return count

# Edge TTS 음성 목록
EDGE_VOICES = [
    {'id': 'ko-KR-SunHiNeural', 'name': '한국어 - 선희 (여성)'},
    {'id': 'ko-KR-InJoonNeural', 'name': '한국어 - 인준 (남성)'},
    {'id': 'en-US-AriaNeural', 'name': '영어 - 아리아 (여성)'},
    {'id': 'en-US-GuyNeural', 'name': '영어 - 가이 (남성)'},
    {'id': 'ja-JP-NanamiNeural', 'name': '일본어 - 나나미 (여성)'},
    {'id': 'zh-CN-XiaoxiaoNeural', 'name': '중국어 - 샤오샤오 (여성)'}
]

# Edge TTS 속도 매핑
EDGE_RATE_MAP = {"1": "-25%", "2": "-15%", "3": "+0%", "4": "+15%", "5": "+25%"}

# 세션 상태 초기화
if 'file_data' not in st.session_state:
    st.session_state.file_data = None
if 'current_row' not in st.session_state:
    st.session_state.current_row = 2
if 'reading' not in st.session_state:
    st.session_state.reading = False
if 'prev_g_value' not in st.session_state:
    st.session_state.prev_g_value = None
if 'tts_engine' not in st.session_state:
    st.session_state.tts_engine = 'Edge TTS (고품질)'
if 'selected_voice' not in st.session_state:
    st.session_state.selected_voice = 'ko-KR-SunHiNeural'
if 'speed' not in st.session_state:
    st.session_state.speed = '3'
if 'announce_group' not in st.session_state:
    st.session_state.announce_group = True

# 메인 헤더
st.title("🎵 엑셀 음성 리더 - 웹버전")
st.markdown("**고급 TTS 시스템으로 엑셀 데이터를 음성으로 변환하세요**")

# 사이드바 설정
with st.sidebar:
    st.header("⚙️ 설정")
    
    # TTS 엔진 선택
    tts_engine = st.selectbox(
        "TTS 엔진",
        ["브라우저 TTS (권장)", "Edge TTS (고품질)", "웹 TTS (온라인)"],
        help="브라우저 TTS가 가장 안정적입니다. Edge TTS는 서버 문제로 실패할 수 있습니다.",
        key="tts_engine_select"
    )
    st.session_state.tts_engine = tts_engine
    
    # 음성 선택
    if tts_engine == "Edge TTS (고품질)":
        voice_options = {voice['name']: voice['id'] for voice in EDGE_VOICES}
        selected_voice_name = st.selectbox("음성 선택", list(voice_options.keys()))
        st.session_state.selected_voice = voice_options[selected_voice_name]
    elif tts_engine == "웹 TTS (온라인)":
        st.session_state.selected_voice = None
        st.info("웹 TTS는 온라인 서비스를 사용합니다.")
    else:
        st.session_state.selected_voice = None
        st.info("브라우저 TTS는 브라우저에서 제공하는 음성을 사용합니다.")
    
    # 속도 설정
    speed = st.selectbox(
        "속도",
        ["1", "2", "3", "4", "5"],
        index=2,
        format_func=lambda x: {
            "1": "매우 느림", "2": "느림", "3": "보통", 
            "4": "빠름", "5": "매우 빠름"
        }[x],
        key="speed_select"
    )
    st.session_state.speed = speed
    
    # 시작 행 설정
    start_row = st.number_input("시작 행", min_value=2, value=2, step=1)
    
    # 옵션들
    st.subheader("📋 옵션")
    auto_advance = st.checkbox("자동 진행", help="자동으로 다음 행으로 이동합니다.")
    announce_group = st.checkbox("브랜드/묶음 변화 알림", value=True, help="G열 값이 바뀔 때 알림합니다.")
    st.session_state.announce_group = announce_group
    
    if auto_advance:
        auto_interval = st.slider("자동 진행 간격 (초)", 0.1, 5.0, 0.3, 0.1)

# 메인 컨텐츠
col1, col2 = st.columns([2, 1])

with col1:
    # 파일 업로드
    st.subheader("📂 파일 업로드")
    uploaded_file = st.file_uploader(
        "엑셀 파일을 선택하세요",
        type=['xlsx', 'xls'],
        help="엑셀 파일(.xlsx, .xls)을 업로드하세요"
    )
    
    if uploaded_file is not None:
        try:
            # 엑셀 파일 읽기
            wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
            ws = wb.active
            
            # 데이터 추출
            data = []
            max_row = ws.max_row
            
            for row in range(2, max_row + 1):  # 헤더 제외
                row_data = {
                    'row': row,
                    'g': str(ws.cell(row=row, column=7).value or "").strip(),
                    'h': str(ws.cell(row=row, column=8).value or "").strip(),
                    'i': str(ws.cell(row=row, column=9).value or "").strip(),
                    'j': str(ws.cell(row=row, column=10).value or "").strip(),
                    'k': str(ws.cell(row=row, column=11).value or "").strip(),
                    'l': ws.cell(row=row, column=12).value
                }
                data.append(row_data)
            
            st.session_state.file_data = {
                'data': data,
                'max_row': max_row,
                'filename': uploaded_file.name
            }
            
            st.success(f"✅ {uploaded_file.name} 파일이 성공적으로 로드되었습니다! ({len(data)}개 행)")
            
        except Exception as e:
            st.error(f"❌ 파일 읽기 실패: {str(e)}")
    
    # 현재 행 표시
    if st.session_state.file_data:
        st.subheader("📊 현재 상태")
        
        # 진행률 표시
        progress = st.session_state.current_row / st.session_state.file_data['max_row']
        st.progress(progress, text=f"진행률: {st.session_state.current_row}/{st.session_state.file_data['max_row']} 행")
        
        # 현재 행 데이터 표시
        if st.session_state.current_row <= len(st.session_state.file_data['data']):
            current_data = st.session_state.file_data['data'][st.session_state.current_row - 2]
            
            st.markdown("**현재 행 데이터:**")
            col_a, col_b, col_c, col_d = st.columns(4)
            
            with col_a:
                st.metric("G열 (브랜드)", current_data['g'] or "없음")
            with col_b:
                st.metric("H열", current_data['h'] or "없음")
            with col_c:
                st.metric("I열 (상품명)", current_data['i'] or "없음")
            with col_d:
                st.metric("J열 (색상)", current_data['j'] or "없음")
            
            col_e, col_f = st.columns(2)
            with col_e:
                st.metric("K열 (사이즈)", current_data['k'] or "없음")
            with col_f:
                st.metric("L열 (수량)", current_data['l'] or "없음")

with col2:
    # 제어 버튼들
    st.subheader("🎮 제어")
    
    col_btn1, col_btn2 = st.columns(2)
    
    with col_btn1:
        if st.button("▶️ 시작", type="primary", use_container_width=True):
            if st.session_state.file_data:
                st.session_state.reading = True
                st.session_state.current_row = start_row
                st.session_state.prev_g_value = None
                st.rerun()
            else:
                st.warning("먼저 파일을 업로드하세요.")
    
    with col_btn2:
        if st.button("⏹️ 중지", type="secondary", use_container_width=True):
            st.session_state.reading = False
            st.rerun()
    
    # 단일 행 읽기
    if st.button("🔊 현재 행 읽기", use_container_width=True):
        if st.session_state.file_data and st.session_state.current_row <= len(st.session_state.file_data['data']):
            # 현재 행 읽기 로직을 직접 구현
            current_data = st.session_state.file_data['data'][st.session_state.current_row - 2]
            
            # 읽을 텍스트 구성
            text_parts = []
            
            # G열 (브랜드/묶음) 처리
            g_value = clean_g_value(current_data['g'])
            if st.session_state.announce_group and g_value and g_value != st.session_state.prev_g_value:
                # 연속된 G값 개수 계산 (간단화)
                g_count = 1
                for i in range(st.session_state.current_row, st.session_state.file_data['max_row'] + 1):
                    if i < len(st.session_state.file_data['data']) + 1:
                        next_data = st.session_state.file_data['data'][i - 2]
                        next_g = clean_g_value(next_data['g'])
                        if next_g == g_value:
                            g_count += 1
                        else:
                            break
                
                if g_count > 1:
                    text_parts.append(f"{g_value} {convert_quantity(g_count)}")
                else:
                    text_parts.append(g_value)
            
            st.session_state.prev_g_value = g_value
            
            # 상품명, 색상, 사이즈, 수량
            if current_data['i']:
                text_parts.append(current_data['i'])
            if current_data['j']:
                text_parts.append(current_data['j'])
            if current_data['k']:
                text_parts.append(convert_size(current_data['k']))
            if current_data['l'] and isinstance(current_data['l'], (int, float)) and current_data['l'] >= 2:
                text_parts.append(convert_quantity(int(current_data['l'])))
            
            combined_text = " ".join(text_parts)
            
            if combined_text.strip():
                st.info(f"🔊 읽을 내용: {combined_text}")
                
                # TTS 엔진 선택
                if st.session_state.tts_engine == "웹 TTS (온라인)":
                    # 웹 TTS 사용 (Google Translate TTS)
                    try:
                        import requests
                        import urllib.parse
                        
                        # Google Translate TTS API 사용
                        text_encoded = urllib.parse.quote(combined_text)
                        tts_url = f"https://translate.google.com/translate_tts?ie=UTF-8&q={text_encoded}&tl=ko&client=tw-ob"
                        
                        response = requests.get(tts_url)
                        if response.status_code == 200:
                            st.audio(response.content, format='audio/mpeg')
                            st.success("웹 TTS로 음성을 재생했습니다.")
                        else:
                            st.error("웹 TTS 서비스에 접근할 수 없습니다.")
                    except Exception as e:
                        st.error(f"웹 TTS 실패: {str(e)}")
                        st.info("브라우저 TTS로 자동 전환합니다...")
                        
                        # 웹 TTS 실패 시 브라우저 TTS로 폴백
                        try:
                            import pyttsx3
                            engine = pyttsx3.init()
                            engine.setProperty('rate', 200)
                            engine.setProperty('volume', 1.0)
                            engine.say(combined_text)
                            engine.runAndWait()
                            st.success("브라우저 TTS로 음성을 재생했습니다.")
                        except Exception as e2:
                            st.error(f"브라우저 TTS도 실패: {str(e2)}")
                            st.info("텍스트를 복사하여 다른 TTS 도구를 사용하세요.")
                elif st.session_state.tts_engine == "Edge TTS (고품질)":
                    # Edge TTS 시도
                    try:
                        rate = EDGE_RATE_MAP.get(st.session_state.speed, "+0%")
                        selected_voice = st.session_state.selected_voice
                        
                        # Edge TTS로 음성 생성
                        with st.spinner("음성을 생성하고 있습니다..."):
                            temp_audio = os.path.join(tempfile.gettempdir(), f"tts_{uuid.uuid4().hex}.mp3")
                            comm = edge_tts.Communicate(combined_text, voice=selected_voice, rate=rate)
                            asyncio.run(comm.save(temp_audio))
                            
                            # 오디오 파일 재생
                            with open(temp_audio, 'rb') as audio_file:
                                audio_bytes = audio_file.read()
                                st.audio(audio_bytes, format='audio/mp3')
                            
                            # 임시 파일 삭제
                            try:
                                os.remove(temp_audio)
                            except:
                                pass
                                
                    except Exception as e:
                        st.warning(f"Edge TTS 실패: {str(e)}")
                        st.info("브라우저 TTS로 자동 전환합니다...")
                        
                        # Edge TTS 실패 시 브라우저 TTS로 폴백
                        try:
                            import pyttsx3
                            engine = pyttsx3.init()
                            engine.setProperty('rate', 200)
                            engine.setProperty('volume', 1.0)
                            engine.say(combined_text)
                            engine.runAndWait()
                            st.success("브라우저 TTS로 음성을 재생했습니다.")
                        except Exception as e2:
                            st.error(f"브라우저 TTS도 실패: {str(e2)}")
                            st.info("텍스트를 복사하여 다른 TTS 도구를 사용하세요.")
                else:
                    # 브라우저 TTS 사용
                    try:
                        import pyttsx3
                        engine = pyttsx3.init()
                        engine.setProperty('rate', 200)
                        engine.setProperty('volume', 1.0)
                        engine.say(combined_text)
                        engine.runAndWait()
                        st.success("브라우저 TTS로 음성을 재생했습니다.")
                    except Exception as e:
                        st.error(f"브라우저 TTS 실패: {str(e)}")
                        st.info("텍스트를 복사하여 다른 TTS 도구를 사용하세요.")
            else:
                st.warning("읽을 내용이 없습니다.")
        else:
            st.warning("읽을 데이터가 없습니다.")
    
    # 다음 행
    if st.button("➡️ 다음 행", use_container_width=True):
        if st.session_state.file_data:
            st.session_state.current_row += 1
            if st.session_state.current_row > st.session_state.file_data['max_row']:
                st.session_state.current_row = st.session_state.file_data['max_row']
                st.info("마지막 행입니다.")
            st.rerun()
    
    # 이전 행
    if st.button("⬅️ 이전 행", use_container_width=True):
        if st.session_state.current_row > 2:
            st.session_state.current_row -= 1
            st.rerun()
    
    # 네이버 검색
    if st.button("🔍 네이버 검색", use_container_width=True):
        if st.session_state.file_data and st.session_state.current_row <= len(st.session_state.file_data['data']):
            current_data = st.session_state.file_data['data'][st.session_state.current_row - 2]
            h_value = current_data['h']
            i_value = current_data['i']
            
            if h_value and i_value:
                query = f"{h_value} {i_value}"
                naver_url = f"https://search.naver.com/search.naver?query={query}"
                
                # JavaScript로 새창 열기
                st.markdown(f"""
                <script>
                window.open('{naver_url}', '_blank');
                </script>
                """, unsafe_allow_html=True)
                
                st.success(f"🔍 '{query}' 검색을 새창에서 열었습니다.")
            else:
                st.warning("검색할 내용이 없습니다.")
    
    # 키보드 단축키 버튼들
    st.subheader("⌨️ 키보드 단축키")
    
    col_kb1, col_kb2, col_kb3 = st.columns(3)
    
    with col_kb1:
        if st.button("1️⃣ 다음 행", use_container_width=True, key="kb_1"):
            if st.session_state.file_data:
                st.session_state.current_row += 1
                if st.session_state.current_row > st.session_state.file_data['max_row']:
                    st.session_state.current_row = st.session_state.file_data['max_row']
                    st.info("마지막 행입니다.")
                st.rerun()
    
    with col_kb2:
        if st.button("2️⃣ 다시 읽기", use_container_width=True, key="kb_2"):
            if st.session_state.file_data:
                # 현재 행 읽기 로직을 직접 구현
                current_data = st.session_state.file_data['data'][st.session_state.current_row - 2]
                
                # 읽을 텍스트 구성
                text_parts = []
                
                # G열 (브랜드/묶음) 처리
                g_value = clean_g_value(current_data['g'])
                if st.session_state.announce_group and g_value and g_value != st.session_state.prev_g_value:
                    # 연속된 G값 개수 계산 (간단화)
                    g_count = 1
                    for i in range(st.session_state.current_row, st.session_state.file_data['max_row'] + 1):
                        if i < len(st.session_state.file_data['data']) + 1:
                            next_data = st.session_state.file_data['data'][i - 2]
                            next_g = clean_g_value(next_data['g'])
                            if next_g == g_value:
                                g_count += 1
                            else:
                                break
                    
                    if g_count > 1:
                        text_parts.append(f"{g_value} {convert_quantity(g_count)}")
                    else:
                        text_parts.append(g_value)
                
                st.session_state.prev_g_value = g_value
                
                # 상품명, 색상, 사이즈, 수량
                if current_data['i']:
                    text_parts.append(current_data['i'])
                if current_data['j']:
                    text_parts.append(current_data['j'])
                if current_data['k']:
                    text_parts.append(convert_size(current_data['k']))
                if current_data['l'] and isinstance(current_data['l'], (int, float)) and current_data['l'] >= 2:
                    text_parts.append(convert_quantity(int(current_data['l'])))
                
                combined_text = " ".join(text_parts)
                
                if combined_text.strip():
                    st.info(f"🔊 읽을 내용: {combined_text}")
                    
                    # TTS 엔진 선택
                    if st.session_state.tts_engine == "웹 TTS (온라인)":
                        # 웹 TTS 사용 (Google Translate TTS)
                        try:
                            import requests
                            import urllib.parse
                            
                            # Google Translate TTS API 사용
                            text_encoded = urllib.parse.quote(combined_text)
                            tts_url = f"https://translate.google.com/translate_tts?ie=UTF-8&q={text_encoded}&tl=ko&client=tw-ob"
                            
                            response = requests.get(tts_url)
                            if response.status_code == 200:
                                st.audio(response.content, format='audio/mpeg')
                                st.success("웹 TTS로 음성을 재생했습니다.")
                            else:
                                st.error("웹 TTS 서비스에 접근할 수 없습니다.")
                        except Exception as e:
                            st.error(f"웹 TTS 실패: {str(e)}")
                            st.info("브라우저 TTS로 자동 전환합니다...")
                            
                            # 웹 TTS 실패 시 브라우저 TTS로 폴백
                            try:
                                import pyttsx3
                                engine = pyttsx3.init()
                                engine.setProperty('rate', 200)
                                engine.setProperty('volume', 1.0)
                                engine.say(combined_text)
                                engine.runAndWait()
                                st.success("브라우저 TTS로 음성을 재생했습니다.")
                            except Exception as e2:
                                st.error(f"브라우저 TTS도 실패: {str(e2)}")
                                st.info("텍스트를 복사하여 다른 TTS 도구를 사용하세요.")
                    elif st.session_state.tts_engine == "Edge TTS (고품질)":
                        # Edge TTS 시도
                        try:
                            rate = EDGE_RATE_MAP.get(st.session_state.speed, "+0%")
                            selected_voice = st.session_state.selected_voice
                            
                            # Edge TTS로 음성 생성
                            with st.spinner("음성을 생성하고 있습니다..."):
                                temp_audio = os.path.join(tempfile.gettempdir(), f"tts_{uuid.uuid4().hex}.mp3")
                                comm = edge_tts.Communicate(combined_text, voice=selected_voice, rate=rate)
                                asyncio.run(comm.save(temp_audio))
                                
                                # 오디오 파일 재생
                                with open(temp_audio, 'rb') as audio_file:
                                    audio_bytes = audio_file.read()
                                    st.audio(audio_bytes, format='audio/mp3')
                                
                                # 임시 파일 삭제
                                try:
                                    os.remove(temp_audio)
                                except:
                                    pass
                                    
                        except Exception as e:
                            st.warning(f"Edge TTS 실패: {str(e)}")
                            st.info("브라우저 TTS로 자동 전환합니다...")
                            
                            # Edge TTS 실패 시 브라우저 TTS로 폴백
                            try:
                                import pyttsx3
                                engine = pyttsx3.init()
                                engine.setProperty('rate', 200)
                                engine.setProperty('volume', 1.0)
                                engine.say(combined_text)
                                engine.runAndWait()
                                st.success("브라우저 TTS로 음성을 재생했습니다.")
                            except Exception as e2:
                                st.error(f"브라우저 TTS도 실패: {str(e2)}")
                                st.info("텍스트를 복사하여 다른 TTS 도구를 사용하세요.")
                    else:
                        # 브라우저 TTS 사용
                        try:
                            import pyttsx3
                            engine = pyttsx3.init()
                            engine.setProperty('rate', 200)
                            engine.setProperty('volume', 1.0)
                            engine.say(combined_text)
                            engine.runAndWait()
                            st.success("브라우저 TTS로 음성을 재생했습니다.")
                        except Exception as e:
                            st.error(f"브라우저 TTS 실패: {str(e)}")
                            st.info("텍스트를 복사하여 다른 TTS 도구를 사용하세요.")
                else:
                    st.warning("읽을 내용이 없습니다.")
    
    with col_kb3:
        if st.button("3️⃣ 네이버 검색", use_container_width=True, key="kb_3"):
            if st.session_state.file_data and st.session_state.current_row <= len(st.session_state.file_data['data']):
                current_data = st.session_state.file_data['data'][st.session_state.current_row - 2]
                h_value = current_data['h']
                i_value = current_data['i']
                
                if h_value and i_value:
                    query = f"{h_value} {i_value}"
                    naver_url = f"https://search.naver.com/search.naver?query={query}"
                    
                    # JavaScript로 새창 열기
                    st.markdown(f"""
                    <script>
                    window.open('{naver_url}', '_blank');
                    </script>
                    """, unsafe_allow_html=True)
                    
                    st.success(f"🔍 '{query}' 검색을 새창에서 열었습니다.")
                else:
                    st.warning("검색할 내용이 없습니다.")

# 엑셀 데이터 미리보기
if st.session_state.file_data:
    st.subheader("📋 엑셀 데이터 미리보기")
    
    # 데이터프레임 생성
    df_data = []
    for item in st.session_state.file_data['data']:
        df_data.append({
            '행': item['row'],
            'G열(브랜드)': item['g'],
            'H열': item['h'],
            'I열(상품명)': item['i'],
            'J열(색상)': item['j'],
            'K열(사이즈)': item['k'],
            'L열(수량)': item['l']
        })
    
    df = pd.DataFrame(df_data)
    
    # 현재 행 하이라이트를 위한 스타일링
    def highlight_current_row(row):
        if row['행'] == st.session_state.current_row:
            return ['background-color: #ffeb3b'] * len(row)
        return [''] * len(row)
    
    # 스타일 적용
    styled_df = df.style.apply(highlight_current_row, axis=1)
    
    # 데이터프레임 표시
    st.dataframe(
        styled_df,
        use_container_width=True,
        height=400,
        hide_index=True
    )
    
    # 현재 행 정보
    st.markdown(f"**현재 선택된 행: {st.session_state.current_row}**")
    
    # 행 이동 버튼들
    col_nav1, col_nav2, col_nav3, col_nav4 = st.columns(4)
    
    with col_nav1:
        if st.button("⏮️ 처음으로", use_container_width=True):
            st.session_state.current_row = 2
            st.rerun()
    
    with col_nav2:
        if st.button("⬅️ 이전 행", use_container_width=True):
            if st.session_state.current_row > 2:
                st.session_state.current_row -= 1
                st.rerun()
    
    with col_nav3:
        if st.button("➡️ 다음 행", use_container_width=True):
            if st.session_state.current_row < st.session_state.file_data['max_row']:
                st.session_state.current_row += 1
                st.rerun()
    
    with col_nav4:
        if st.button("⏭️ 마지막으로", use_container_width=True):
            st.session_state.current_row = st.session_state.file_data['max_row']
            st.rerun()



# 자동 진행 처리
if st.session_state.reading and auto_advance:
    if st.session_state.current_row < st.session_state.file_data['max_row']:
        st.session_state.current_row += 1
        # 자동 진행 시에는 음성 재생하지 않고 행만 이동
        st.rerun()
    else:
        st.session_state.reading = False
        st.success("모든 행을 읽었습니다!")

# 사용법 안내
st.markdown("---")
st.markdown("### 🎯 사용법")
st.markdown("""
1. **파일 업로드**: 엑셀 파일을 업로드하세요
2. **설정 조정**: 사이드바에서 TTS 설정을 조정하세요
3. **읽기 시작**: '시작' 버튼을 클릭하거나 '현재 행 읽기'를 사용하세요
4. **자동 진행**: 필요시 자동 진행을 활성화하세요
5. **키보드 단축키**: 1(다음 행), 2(다시 읽기), 3(네이버 검색) 버튼을 사용하세요
""")