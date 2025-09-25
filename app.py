from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
import tempfile
import os
import uuid
import asyncio
import edge_tts
from werkzeug.utils import secure_filename
import json
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 전역 변수
current_row = 2
file_data = None
reading = False
tts_queue = []

# TTS 엔진 설정
TTS_ENGINES = {
    'edge-tts': {
        'name': 'Edge TTS',
        'voices': [
            {'id': 'ko-KR-SunHiNeural', 'name': '한국어 - 선희 (여성)'},
            {'id': 'ko-KR-InJoonNeural', 'name': '한국어 - 인준 (남성)'},
            {'id': 'en-US-AriaNeural', 'name': '영어 - 아리아 (여성)'},
            {'id': 'en-US-GuyNeural', 'name': '영어 - 가이 (남성)'},
            {'id': 'ja-JP-NanamiNeural', 'name': '일본어 - 나나미 (여성)'},
            {'id': 'zh-CN-XiaoxiaoNeural', 'name': '중국어 - 샤오샤오 (여성)'}
        ]
    },
    'browser-tts': {
        'name': '브라우저 TTS',
        'voices': 'browser_voices'  # 브라우저에서 제공하는 음성들
    }
}

# 사이즈 변환 딕셔너리
SIZE_DICT = {
    "XS": "엑스스몰", "S": "스몰", "M": "미디움", "L": "라지", "FREE": "프리",
    "XL": "엑스라지", "XXL": "투엑스라지",
    "JS": "주니어 스몰", "JM": "주니어 미디움", "JL": "주니어 라지"
}

# 한국어 숫자 변환
KOREAN_NUMBER_MAP = {
    1: "한개", 2: "두개", 3: "세개", 4: "네개",
    5: "다섯개", 6: "여섯개", 7: "일곱개",
    8: "여덟개", 9: "아홉개", 10: "열개"
}

def convert_size(size_code):
    return SIZE_DICT.get(str(size_code).upper(), size_code)

def convert_quantity(n):
    if isinstance(n, int) and n >= 1:
        return KOREAN_NUMBER_MAP.get(n, f"{n}개")
    return ""

def clean_g_value(text):
    import re
    return re.sub(r"\(.*?\)", "", text).strip()

def count_consecutive_g_values(ws, start_row):
    current_g = clean_g_value(str(ws.cell(row=start_row, column=7).value or ""))
    count = 0
    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        g_value = clean_g_value(str(ws.cell(row=row, column=7).value or ""))
        if g_value == current_g:
            count += 1
        else:
            break
    return count

@app.route('/')
def index():
    return render_template('index.html', tts_engines=TTS_ENGINES)

@app.route('/upload', methods=['POST'])
def upload_file():
    global file_data, current_row
    
    if 'file' not in request.files:
        return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400
    
    if file and file.filename.lower().endswith(('.xlsx', '.xls')):
        try:
            # 임시 파일로 저장
            temp_path = os.path.join(tempfile.gettempdir(), f"temp_{uuid.uuid4().hex}.xlsx")
            file.save(temp_path)
            
            # 엑셀 파일 읽기
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            ws = wb.active
            
            # 데이터 추출
            data = []
            max_row = ws.max_row
            
            for row in range(2, max_row + 1):  # 헤더 제외
                row_data = {
                    'row': row,
                    'g': str(ws.cell(row=row, column=7).value or "").strip(),
                    'h': str(ws.cell(row=row, column=8).value or "").strip(),
                    'i': str(ws.cell(row=row, column=9).value or "").strip(),
                    'j': str(ws.cell(row=row, column=10).value or "").strip(),
                    'k': str(ws.cell(row=row, column=11).value or "").strip(),
                    'l': ws.cell(row=row, column=12).value
                }
                data.append(row_data)
            
            file_data = {
                'path': temp_path,
                'data': data,
                'max_row': max_row
            }
            
            current_row = 2
            return jsonify({
                'success': True,
                'filename': file.filename,
                'total_rows': len(data),
                'message': f'{len(data)}개의 행을 읽었습니다.'
            })
            
        except Exception as e:
            return jsonify({'error': f'파일 읽기 실패: {str(e)}'}), 500
    else:
        return jsonify({'error': '엑셀 파일만 업로드 가능합니다.'}), 400

@app.route('/read_row', methods=['POST'])
def read_row():
    global current_row, reading
    
    if not file_data:
        return jsonify({'error': '파일을 먼저 업로드하세요.'}), 400
    
    data = request.get_json()
    row_num = data.get('row', current_row)
    engine = data.get('engine', 'edge-tts')
    voice = data.get('voice', 'ko-KR-SunHiNeural')
    speed = data.get('speed', 1.0)
    
    try:
        # 해당 행 데이터 가져오기
        row_data = None
        for data_row in file_data['data']:
            if data_row['row'] == row_num:
                row_data = data_row
                break
        
        if not row_data:
            return jsonify({'error': '해당 행을 찾을 수 없습니다.'}), 404
        
        # 읽을 텍스트 구성
        text_parts = []
        
        # G열 (브랜드/묶음) 처리
        g_value = clean_g_value(row_data['g'])
        if g_value:
            # 연속된 G값 개수 계산
            wb = openpyxl.load_workbook(file_data['path'], data_only=True)
            ws = wb.active
            g_count = count_consecutive_g_values(ws, row_num)
            
            if g_count > 1:
                text_parts.append(f"{g_value} {convert_quantity(g_count)}")
            else:
                text_parts.append(g_value)
        
        # 상품명, 색상, 사이즈, 수량
        if row_data['i']:
            text_parts.append(row_data['i'])
        if row_data['j']:
            text_parts.append(row_data['j'])
        if row_data['k']:
            text_parts.append(convert_size(row_data['k']))
        if row_data['l'] and isinstance(row_data['l'], (int, float)) and row_data['l'] >= 2:
            text_parts.append(convert_quantity(int(row_data['l'])))
        
        combined_text = " ".join(text_parts)
        
        if not combined_text.strip():
            return jsonify({'error': '읽을 내용이 없습니다.'}), 400
        
        # TTS 처리
        if engine == 'edge-tts':
            # Edge TTS로 음성 생성
            rate_map = {"1": "-25%", "2": "-15%", "3": "+0%", "4": "+15%", "5": "+25%"}
            rate = rate_map.get(str(int(speed)), "+0%")
            
            # 비동기 TTS 처리
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            try:
                temp_audio = os.path.join(tempfile.gettempdir(), f"tts_{uuid.uuid4().hex}.mp3")
                comm = edge_tts.Communicate(combined_text, voice=voice, rate=rate)
                loop.run_until_complete(comm.save(temp_audio))
                
                return jsonify({
                    'success': True,
                    'text': combined_text,
                    'audio_url': f'/audio/{os.path.basename(temp_audio)}',
                    'row_data': row_data
                })
            finally:
                loop.close()
        else:
            # 브라우저 TTS 사용
            return jsonify({
                'success': True,
                'text': combined_text,
                'use_browser_tts': True,
                'row_data': row_data
            })
            
    except Exception as e:
        return jsonify({'error': f'TTS 처리 실패: {str(e)}'}), 500

@app.route('/audio/<filename>')
def serve_audio(filename):
    try:
        audio_path = os.path.join(tempfile.gettempdir(), filename)
        if os.path.exists(audio_path):
            return send_file(audio_path, as_attachment=False)
        else:
            return jsonify({'error': '오디오 파일을 찾을 수 없습니다.'}), 404
    except Exception as e:
        return jsonify({'error': f'오디오 파일 서빙 실패: {str(e)}'}), 500

@app.route('/search_naver', methods=['POST'])
def search_naver():
    if not file_data:
        return jsonify({'error': '파일을 먼저 업로드하세요.'}), 400
    
    data = request.get_json()
    row_num = data.get('row', current_row)
    
    try:
        # 해당 행 데이터 찾기
        row_data = None
        for data_row in file_data['data']:
            if data_row['row'] == row_num:
                row_data = data_row
                break
        
        if not row_data:
            return jsonify({'error': '해당 행을 찾을 수 없습니다.'}), 404
        
        h_value = row_data['h']
        i_value = row_data['i']
        
        if h_value and i_value:
            query = f"{h_value} {i_value}"
            naver_url = f"https://search.naver.com/search.naver?query={query}"
            return jsonify({
                'success': True,
                'url': naver_url,
                'query': query
            })
        else:
            return jsonify({'error': '검색할 내용이 없습니다.'}), 400
            
    except Exception as e:
        return jsonify({'error': f'검색 실패: {str(e)}'}), 500

@app.route('/get_voices')
def get_voices():
    return jsonify(TTS_ENGINES)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
