import os
import openpyxl
import threading
import re
import webbrowser
import tempfile
import uuid
from queue import Queue
from tkinter import (
    Tk, Button, Frame, filedialog,
    Entry, Label, messagebox, StringVar, OptionMenu, Checkbutton, BooleanVar
)

# ====== Tk 루트는 최우선으로 생성 ======
root = Tk()
root.title("엑셀 음성 리더 (edge-tts 고속모드)")
root.geometry("920x600")
root.configure(padx=20, pady=20)

# ====== 새로 추가된 모듈 ======
import asyncio
import edge_tts
from playsound import playsound

# ====== 폴백 엔진 ======
try:
    import pyttsx3
    PYTTSX3_AVAILABLE = True
except Exception:
    PYTTSX3_AVAILABLE = False

# =========================
# 전역 상태
# =========================
current_row = 2
file_path = None
reading = False

tts_queue = Queue()
last_spoken = [None]

engine_mode = StringVar(master=root, value="edge-tts")  # "edge-tts" | "pyttsx3"
auto_advance_var = BooleanVar(master=root, value=False)
speed_var = StringVar(master=root, value="3")  # 1~5
# ✅ G열 이름 읽기 기본 ON (예전 방식 유지)
announce_group_var = BooleanVar(master=root, value=True)

EDGE_RATE_MAP = {"1": "-25%", "2": "-15%", "3": "+0%", "4": "+15%", "5": "+25%"}
PYTTS_RATE_MAP = {"1": 150, "2": 175, "3": 200, "4": 225, "5": 260}

_pyttsx_engine = [None]
def _pytts_engine_ref():
    return _pytts_engine[0]

def _init_pyttsx():
    if not PYTTSX3_AVAILABLE:
        return None
    try:
        eng = pyttsx3.init(driverName='sapi5')
        rate = PYTTS_RATE_MAP.get(speed_var.get(), 200)
        eng.setProperty('rate', rate)
        eng.setProperty('volume', 1.0)
        _pytts_engine[0] = eng
        return eng
    except Exception:
        return None

# =========================
# 유틸
# =========================
def set_speed(sp):
    speed_var.set(str(sp))

def open_excel():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        try:
            os.startfile(file_path)
            file_label.config(text=f"선택된 파일: {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("오류", f"파일 열기 실패: {e}")

def update_display_text(row_num, i_value, j_value, k_value, l_value):
    # 행번호는 화면 표시만, 음성으로는 읽지 않음
    display_text.set(f"{row_num} | {i_value}  {j_value}  {k_value}  {l_value}")

def clean_g_value(text):
    return re.sub(r"\(.*?\)", "", text).strip()

def count_consecutive_g_values(ws, start_row):
    current_g = clean_g_value(str(ws.cell(row=start_row, column=7).value or ""))
    count = 0
    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        g_value = clean_g_value(str(ws.cell(row=row, column=7).value or ""))
        if g_value == current_g:
            count += 1
        else:
            break
    return count

size_dict = {
    "XS": "엑스스몰", "S": "스몰", "M": "미디움", "L": "라지", "FREE": "프리",
    "XL": "엑스라지", "XXL": "투엑스라지",
    "JS": "주니어 스몰", "JM": "주니어 미디움", "JL": "주니어 라지"
}
def convert_size(size_code):
    return size_dict.get(str(size_code).upper(), size_code)

korean_number_map = {
    1: "한개", 2: "두개", 3: "세개", 4: "네개",
    5: "다섯개", 6: "여섯개", 7: "일곱개",
    8: "여덟개", 9: "아홉개", 10: "열개"
}
def convert_quantity(n):
    if isinstance(n, int) and n >= 1:
        return korean_number_map.get(n, f"{n}개")
    return ""

# =========================
# 발화 큐잉 (한 행 = 한 문장 합성)
# =========================
def speak(text, force=False):
    if not text:
        return
    if force or text != last_spoken[0]:
        tts_queue.put(("SAY", text))
        last_spoken[0] = text

def flush_tts_queue():
    try:
        while not tts_queue.empty():
            _ = tts_queue.get_nowait()
            tts_queue.task_done()
    except Exception:
        pass

# =========================
# TTS 재생(엔진별)
# =========================
async def _edge_tts_synthesize_to_mp3(text, voice="ko-KR-SunHiNeural", rate="+0%"):
    temp_dir = tempfile.gettempdir()
    out_path = os.path.join(temp_dir, f"tts_{uuid.uuid4().hex}.mp3")
    comm = edge_tts.Communicate(text, voice=voice, rate=rate)
    await comm.save(out_path)
    return out_path

def _play_mp3_blocking(mp3_path):
    playsound(mp3_path)

def _edge_tts_say(text):
    rate = EDGE_RATE_MAP.get(speed_var.get(), "+0%")
    voice = "ko-KR-SunHiNeural"
    mp3_path = asyncio.run(_edge_tts_synthesize_to_mp3(text, voice=voice, rate=rate))
    try:
        _play_mp3_blocking(mp3_path)
    finally:
        try:
            os.remove(mp3_path)
        except Exception:
            pass

def _pyttsx3_say(text):
    eng = _pytts_engine_ref() or _init_pyttsx()
    if not eng:
        raise RuntimeError("pyttsx3 초기화 실패")
    rate = PYTTS_RATE_MAP.get(speed_var.get(), 200)
    eng.setProperty('rate', rate)
    eng.setProperty('volume', 1.0)
    eng.say(text)
    eng.runAndWait()

# =========================
# TTS 워커(단일 스레드)
# =========================
_stop_worker = [False]
def tts_worker():
    while not _stop_worker[0]:
        cmd = tts_queue.get()
        if cmd is None:
            tts_queue.task_done()
            break
        typ, payload = cmd
        if typ == "SAY":
            text = payload
            try:
                if engine_mode.get() == "edge-tts":
                    _edge_tts_say(text)
                else:
                    _pyttsx3_say(text)
            except Exception:
                # edge-tts 실패시 pyttsx3 폴백
                if engine_mode.get() == "edge-tts" and PYTTSX3_AVAILABLE:
                    try:
                        _pyttsx3_say(text)
                    except Exception:
                        pass
            finally:
                tts_queue.task_done()
        else:
            tts_queue.task_done()

tts_thread = threading.Thread(target=tts_worker, daemon=True)
tts_thread.start()

# =========================
# 핵심 읽기 로직 (행번호 미발화 + G열 이름 복구)
# =========================
prev_g_value = None
def read_current_row(force_read=False):
    global current_row, prev_g_value, reading

    if not file_path:
        messagebox.showwarning("경고", "엑셀 파일을 먼저 열어주세요")
        return False

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("오류", f"엑셀 로드 실패: {e}\n파일이 열려 있으면 닫아주세요.")
        return False

    max_row = ws.max_row
    if current_row > max_row:
        messagebox.showinfo("알림", "더 이상 읽을 행이 없습니다")
        reading = False
        return False

    g_value = clean_g_value(str(ws.cell(row=current_row, column=7).value or ""))
    h_value = str(ws.cell(row=current_row, column=8).value or "").strip()
    i_value = str(ws.cell(row=current_row, column=9).value or "").strip()
    j_value = str(ws.cell(row=current_row, column=10).value or "").strip()
    k_value = str(ws.cell(row=current_row, column=11).value or "").strip()

    # L열 수량: 2 이상만 읽기
    try:
        l_raw_value = ws.cell(row=current_row, column=12).value
        l_value = int(l_raw_value) if l_raw_value and str(l_raw_value).isdigit() else None
        if l_value is not None and l_value < 2:
            l_value = None
    except (ValueError, TypeError):
        l_value = None

    display_l_value = str(l_value) if l_value else ""
    update_display_text(current_row, i_value, j_value, k_value, display_l_value)

    # ===== 한 행 = 한 문장으로 합성 =====
    parts = []

    # ✅ G열(브랜드/묶음) 읽기: 값이 바뀌면 예전 방식대로 읽기(연속 개수 포함)
    if announce_group_var.get() and g_value and g_value != prev_g_value:
        g_count = count_consecutive_g_values(ws, current_row)
        if g_count > 1:
            parts.append(f"{g_value} {convert_quantity(g_count)}")
        else:
            parts.append(g_value)
    prev_g_value = g_value

    # 필수: 상품명, 색상, 사이즈, 수량
    if i_value:
        parts.append(i_value)
    if j_value:
        parts.append(j_value)
    if k_value:
        parts.append(convert_size(k_value))
    if l_value is not None:
        qty_txt = convert_quantity(l_value)
        if qty_txt:
            parts.append(qty_txt)

    combined = " ".join([p for p in parts if p])
    if combined.strip():
        speak(combined, force=force_read)

    return True

# =========================
# 진행 제어
# =========================
def start_reading():
    global reading, prev_g_value
    if reading:
        return
    if not file_path:
        messagebox.showwarning("경고", "파일을 먼저 선택하세요")
        return

    prev_g_value = None
    last_spoken[0] = None
    reading = True

    read_current_row(force_read=True)
    if auto_advance_var.get():
        schedule_auto_next()

def stop_reading():
    global reading
    reading = False
    flush_tts_queue()

def schedule_auto_next():
    if not reading or not auto_advance_var.get():
        return
    try:
        # 기본 300ms (필요 시 150~250으로 더 낮출 수 있음)
        interval_ms = int(auto_interval_entry.get()) if auto_interval_entry.get().isdigit() else 300
    except:
        interval_ms = 300
    root.after(interval_ms, auto_next_step)

def auto_next_step():
    if not reading or not auto_advance_var.get():
        return
    next_row()
    schedule_auto_next()

def next_row(event=None):
    global current_row
    if not reading:
        return
    current_row += 1
    read_current_row(force_read=True)

def reread(event=None):
    if not reading:
        return
    read_current_row(force_read=True)

def search_naver(event=None):
    if not file_path:
        messagebox.showwarning("경고", "파일을 먼저 선택하세요")
        return
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        h_value = str(ws.cell(row=current_row, column=8).value or "").strip()
        i_value = str(ws.cell(row=current_row, column=9).value or "").strip()
    except Exception as e:
        messagebox.showerror("오류", f"검색 실패: {e}")
        return
    if h_value and i_value:
        query = f"{h_value} {i_value}"
        webbrowser.open(f"https://search.naver.com/search.naver?query={query}")
    else:
        messagebox.showinfo("알림", "검색할 내용이 없습니다")

def change_start_row():
    global current_row
    try:
        new_row = int(start_row_entry.get())
        if 2 <= new_row:
            current_row = new_row
            messagebox.showinfo("알림", f"{new_row}행으로 변경 완료")
            if reading:
                read_current_row(force_read=True)
        else:
            messagebox.showwarning("경고", "2 이상의 숫자를 입력하세요")
    except ValueError:
        messagebox.showerror("오류", "숫자를 입력해주세요")

def on_toggle_auto():
    if reading and auto_advance_var.get():
        schedule_auto_next()

# =========================
# GUI 레이아웃
# =========================
main_frame = Frame(root)
main_frame.pack(fill='both', expand=True)

left_frame = Frame(main_frame)
left_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=5)
right_frame = Frame(main_frame)
right_frame.grid(row=0, column=1, sticky='nsew', padx=10, pady=5)

main_frame.columnconfigure(0, weight=7)
main_frame.columnconfigure(1, weight=3)

file_frame = Frame(left_frame)
file_frame.pack(fill='x', pady=5)
Button(file_frame, text="📂 파일 열기", command=open_excel, width=15).pack(side='left')
file_label = Label(file_frame, text=" 파일 없음", anchor='w')
file_label.pack(side='left', fill='x', expand=True, padx=10)

row_frame = Frame(left_frame)
row_frame.pack(fill='x', pady=5)
Label(row_frame, text="시작 행:").pack(side='left')
start_row_entry = Entry(row_frame, width=10)
start_row_entry.pack(side='left', padx=5)
Button(row_frame, text="변경", command=change_start_row).pack(side='left')

display_frame = Frame(left_frame, bg='#f0f0f0', padx=15, pady=15)
display_frame.pack(fill='both', expand=True, pady=10)
display_text = StringVar(master=root, value=">> 준비 완료 <<")
Label(display_frame, textvariable=display_text,
      font=('맑은 고딕', 12, 'bold'), bg='#f0f0f0').pack()

control_frame = Frame(left_frame)
control_frame.pack(fill='x', pady=10)
Button(control_frame, text="▶ 시작", command=start_reading,
       bg='#0078D4', fg='white', font=('맑은 고딕', 14), height=2).pack(side='left', fill='x', expand=True, padx=4)
Button(control_frame, text="⏹ 중지", command=stop_reading,
       bg='#FF0000', fg='white', font=('맑은 고딕', 14), height=2).pack(side='left', fill='x', expand=True, padx=4)

# 오른쪽 패널
speed_frame = Frame(right_frame)
speed_frame.pack(pady=10, fill='x')
Label(speed_frame, text="🔊 음성 속도", font=('맑은 고딕', 11)).pack()
OptionMenu(speed_frame, speed_var, "1", "2", "3", "4", "5", command=set_speed).pack()

engine_frame = Frame(right_frame)
engine_frame.pack(pady=10, fill='x')
Label(engine_frame, text="🧠 엔진 선택 (문제 시 pyttsx3로 전환)", font=('맑은 고딕', 11)).pack()
OptionMenu(engine_frame, engine_mode, "edge-tts", "pyttsx3").pack()

group_frame = Frame(right_frame)
group_frame.pack(pady=10, fill='x')
Checkbutton(group_frame, text="브랜드/묶음(G열) 변화 알림", variable=announce_group_var).pack(anchor='w')

shortcut_frame = Frame(right_frame)
shortcut_frame.pack(pady=10, fill='x')
Label(shortcut_frame, text="⌨️ 단축키", font=('맑은 고딕', 11)).pack()
Label(shortcut_frame,
      text="1: 다음 행   2: 현재 행 다시 읽기   3: 네이버 검색\n스페이스: 다음 행",
      justify='left').pack(pady=5)

auto_frame = Frame(right_frame)
auto_frame.pack(pady=10, fill='x')
Checkbutton(auto_frame, text="자동 진행 (고속)", variable=auto_advance_var,
            command=on_toggle_auto).pack(anchor='w')
Label(auto_frame, text="자동 진행 간격 (ms):").pack(anchor='w')
auto_interval_entry = Entry(auto_frame, width=10)
auto_interval_entry.insert(0, "300")  # 기본 300ms
auto_interval_entry.pack(anchor='w')

# 키 바인딩
root.bind('1', lambda e: next_row())
root.bind('<space>', lambda e: next_row())
root.bind('2', lambda e: reread())
root.bind('3', lambda e: search_naver())

def on_closing():
    global reading
    reading = False
    _stop_worker[0] = True
    flush_tts_queue()
    try:
        tts_queue.put(None)
    except Exception:
        pass
    root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()
